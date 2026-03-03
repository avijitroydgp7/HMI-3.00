# project\comment\import_handler.py
"""
Import handler for comment tables - manages import operations from various file formats
(Excel, CSV, JSON) with full formatting preservation.
"""

import csv
import json
import logging

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from PySide6.QtWidgets import QMessageBox, QFileDialog

logger = logging.getLogger(__name__)


class ImportHandler:
    """
    Handles high-level import operations with user dialogs and format detection.
    Works with a Spreadsheet instance to delegate actual file I/O.
    """

    def __init__(self, spreadsheet_widget):
        """
        Initialize the import handler.

        Args:
            spreadsheet_widget: Spreadsheet instance to import into
        """
        self.spreadsheet = spreadsheet_widget

    def import_from_file(self, parent=None):
        """
        Import table data from Excel, CSV, or JSON file.

        Args:
            parent: Parent widget for dialogs
        """
        file_path, _ = QFileDialog.getOpenFileName(
            parent,
            "Import Table Data",
            "",
            "Excel Files (*.xlsx);;CSV Files (*.csv);;JSON Files (*.json);;All Files (*)",
        )
        if not file_path:
            return

        try:
            if file_path.lower().endswith(".xlsx"):
                SpreadsheetImportHandler.import_from_excel(self.spreadsheet, file_path)
            elif file_path.lower().endswith(".json"):
                SpreadsheetImportHandler.import_from_json(self.spreadsheet, file_path)
            else:
                SpreadsheetImportHandler.import_from_csv(self.spreadsheet, file_path)

            self.spreadsheet.save_data_to_service()
            QMessageBox.information(
                parent,
                "Success",
                f"Table imported successfully from:\n{file_path}",
            )
        except Exception as e:
            QMessageBox.critical(
                parent, "Import Error", f"Failed to import table:\n{str(e)}"
            )


class SpreadsheetImportHandler:
    """
    Low-level import operations - performs actual file I/O and data restoration.
    Static methods that work directly with Spreadsheet instances.
    """

    @staticmethod
    def import_from_excel(spreadsheet, file_path):
        """
        Import table data from Excel file with formatting.

        Args:
            spreadsheet: Spreadsheet instance to import into
            file_path: Source file path

        Raises:
            Exception: If openpyxl is not available or import fails
        """
        if not OPENPYXL_AVAILABLE:
            raise Exception(
                "openpyxl library is not installed. Install with: pip install openpyxl"
            )

        try:
            wb = load_workbook(file_path)
            ws = wb.active

            if not ws:
                raise Exception("No worksheet found in Excel file")

            # Determine table dimensions
            max_row = 0
            max_col = 0
            import_data = {}

            # Read all cells with data
            for row_idx, row in enumerate(ws.iter_rows()):
                for col_idx, cell in enumerate(row):
                    if cell.value is None:
                        continue

                    max_row = max(max_row, row_idx)
                    max_col = max(max_col, col_idx)

                    # Extract formatting
                    cell_data = {"value": str(cell.value)}

                    # Font formatting
                    font_data = {}
                    if cell.font:
                        font_data["bold"] = cell.font.bold or False
                        font_data["italic"] = cell.font.italic or False
                        font_data["underline"] = bool(cell.font.underline) or False
                    cell_data["font"] = font_data

                    # Text color
                    text_color = None
                    if cell.font and cell.font.color:
                        color_val = cell.font.color.rgb
                        if color_val and color_val != "00000000":
                            text_color = (
                                f"#{color_val[-6:]}"
                                if len(str(color_val)) >= 6
                                else None
                            )
                    cell_data["text_color"] = text_color

                    # Background color
                    bg_color = None
                    if cell.fill and cell.fill.start_color:
                        color_val = cell.fill.start_color.rgb
                        if color_val and color_val != "00000000":
                            bg_color = (
                                f"#{color_val[-6:]}"
                                if len(str(color_val)) >= 6
                                else None
                            )
                    cell_data["bg_color"] = bg_color

                    import_data[(row_idx, col_idx)] = cell_data

            # Resize table
            if max_row > 0 or max_col > 0:
                spreadsheet.setRowCount(max_row + 1)
                spreadsheet.setColumnCount(max_col + 1)
                spreadsheet.update_headers()

            # Import data - Import SpreadsheetItem here to avoid circular import
            from .comment_table import SpreadsheetItem
            
            spreadsheet.blockSignals(True)
            for (row, col), cell_data in import_data.items():
                item = spreadsheet.item(row, col) or SpreadsheetItem()
                spreadsheet.setItem(row, col, item)
                item.set_data(cell_data)

            spreadsheet.blockSignals(False)
            spreadsheet.evaluate_all_cells()

        except Exception as e:
            raise Exception(f"Failed to import from Excel: {str(e)}")

    @staticmethod
    def import_from_csv(spreadsheet, file_path):
        """
        Import table data from CSV file, preserving formatting.

        Args:
            spreadsheet: Spreadsheet instance to import into
            file_path: Source file path

        Raises:
            Exception: If import fails
        """
        try:
            # Import here to avoid circular imports
            from .comment_table import SpreadsheetItem

            # First, read all the data to determine dimensions
            import_data = {}
            max_row = 0
            max_col = 0

            with open(file_path, "r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f)
                for row_dict in reader:
                    try:
                        row = int(row_dict["Row"])
                        col = int(row_dict["Column"])
                        max_row = max(max_row, row)
                        max_col = max(max_col, col)
                        import_data[(row, col)] = row_dict
                    except (ValueError, KeyError):
                        continue

            # Ensure table is large enough
            if spreadsheet.rowCount() <= max_row:
                spreadsheet.setRowCount(max_row + 1)
            if spreadsheet.columnCount() <= max_col:
                spreadsheet.setColumnCount(max_col + 1)

            spreadsheet.update_headers()

            # Import the data
            spreadsheet.blockSignals(True)
            for (row, col), row_dict in import_data.items():
                item = spreadsheet.item(row, col) or SpreadsheetItem()
                spreadsheet.setItem(row, col, item)

                cell_data = {
                    "value": row_dict.get("Value", ""),
                    "font": {
                        "bold": row_dict.get("FontBold", "").lower()
                        in ("true", "1", "yes"),
                        "italic": row_dict.get("FontItalic", "").lower()
                        in ("true", "1", "yes"),
                        "underline": row_dict.get("FontUnderline", "").lower()
                        in ("true", "1", "yes"),
                    },
                    "text_color": row_dict.get("TextColor", "") or None,
                    "bg_color": row_dict.get("BGColor", "") or None,
                }
                item.set_data(cell_data)

            spreadsheet.blockSignals(False)
            spreadsheet.evaluate_all_cells()

        except Exception as e:
            raise Exception(f"Failed to import from CSV: {str(e)}")

    @staticmethod
    def import_from_json(spreadsheet, file_path):
        """
        Import table data from JSON file with full formatting.

        Args:
            spreadsheet: Spreadsheet instance to import into
            file_path: Source file path

        Raises:
            Exception: If import fails
        """
        try:
            # Import here to avoid circular imports
            from .comment_table import SpreadsheetItem

            with open(file_path, "r", encoding="utf-8") as f:
                data_dict = json.load(f)
                table_data = data_dict.get("table_data", [])

            if not table_data:
                return

            # Resize table to fit imported data
            spreadsheet.setRowCount(len(table_data))
            if table_data:
                spreadsheet.setColumnCount(len(table_data[0]))

            spreadsheet.update_headers()

            # Import the data
            spreadsheet.blockSignals(True)
            for r, row in enumerate(table_data):
                for c, cell_data in enumerate(row):
                    item = spreadsheet.item(r, c) or SpreadsheetItem()
                    spreadsheet.setItem(r, c, item)
                    # Ensure cell_data is a dict with required keys
                    if isinstance(cell_data, dict):
                        item.set_data(cell_data)
                    else:
                        item.set_data({"value": str(cell_data)})

            spreadsheet.blockSignals(False)
            spreadsheet.evaluate_all_cells()

        except Exception as e:
            raise Exception(f"Failed to import from JSON: {str(e)}")
