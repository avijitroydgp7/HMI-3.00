# project\comment\export_handler.py
"""
Export handler for comment tables - manages export operations to various file formats
(Excel, CSV, JSON) with full formatting preservation.
"""

import csv
import json
import logging

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from PySide6.QtWidgets import QMessageBox, QFileDialog

logger = logging.getLogger(__name__)


class ExportHandler:
    """
    Handles high-level export operations with user dialogs and format selection.
    Works with a Spreadsheet instance to delegate actual file I/O.
    """

    def __init__(self, spreadsheet_widget):
        """
        Initialize the export handler.

        Args:
            spreadsheet_widget: Spreadsheet instance to export from
        """
        self.spreadsheet = spreadsheet_widget

    def export_with_format_selection(self, parent=None):
        """
        Export table data with format selection via file save dialog.
        
        Args:
            parent: Parent widget for dialogs
        """
        file_filters = "Excel Files (*.xlsx);;CSV Files (*.csv);;JSON Files (*.json);;All Files (*)"

        file_path, selected_filter = QFileDialog.getSaveFileName(
            parent, "Export Table", "", file_filters
        )

        if not file_path:
            return

        # Determine export format from selected filter or file extension
        if "Excel" in selected_filter or file_path.lower().endswith(".xlsx"):
            self._export_as_excel(file_path, parent)
        elif "CSV" in selected_filter or file_path.lower().endswith(".csv"):
            self._export_as_csv(file_path, parent)
        elif "JSON" in selected_filter or file_path.lower().endswith(".json"):
            self._export_as_json(file_path, parent)
        else:
            # Default to CSV if no clear format detected
            self._export_as_csv(file_path, parent)

    def _export_as_excel(self, file_path, parent=None):
        """
        Export table data to Excel file with native cell formatting.

        Args:
            file_path: Destination file path
            parent: Parent widget for dialogs
        """
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(
                parent,
                "Excel Support Not Available",
                "The openpyxl library is required for Excel export.\n\n"
                "Install it with: pip install openpyxl\n\n"
                "Using CSV export instead.",
            )
            self._export_as_csv(file_path.replace(".xlsx", ".csv"), parent)
            return

        try:
            SpreadsheetExportHandler.export_to_excel(self.spreadsheet, file_path)
            QMessageBox.information(
                parent,
                "Success",
                f"Table exported successfully to:\n{file_path}",
            )
        except Exception as e:
            QMessageBox.critical(
                parent, "Export Error", f"Failed to export table:\n{str(e)}"
            )

    def _export_as_csv(self, file_path, parent=None):
        """
        Export table data to CSV file with formatting.

        Args:
            file_path: Destination file path
            parent: Parent widget for dialogs
        """
        try:
            SpreadsheetExportHandler.export_to_csv(self.spreadsheet, file_path)
            QMessageBox.information(
                parent,
                "Success",
                f"Table exported successfully to:\n{file_path}",
            )
        except Exception as e:
            QMessageBox.critical(
                parent, "Export Error", f"Failed to export table:\n{str(e)}"
            )

    def _export_as_json(self, file_path, parent=None):
        """
        Export table data to JSON file with full formatting.

        Args:
            file_path: Destination file path
            parent: Parent widget for dialogs
        """
        try:
            SpreadsheetExportHandler.export_to_json(self.spreadsheet, file_path)
            QMessageBox.information(
                parent,
                "Success",
                f"Table exported successfully to:\n{file_path}",
            )
        except Exception as e:
            QMessageBox.critical(
                parent, "Export Error", f"Failed to export table:\n{str(e)}"
            )

    def export_to_excel(self, parent=None):
        """
        Export table to Excel with file dialog.

        Args:
            parent: Parent widget for dialogs
        """
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(
                parent,
                "Excel Support Not Available",
                "The openpyxl library is required for Excel export.\n\n"
                "Install it with: pip install openpyxl\n\n"
                "Using CSV export instead.",
            )
            self.export_to_csv(parent)
            return

        file_path, _ = QFileDialog.getSaveFileName(
            parent, "Export Table as Excel", "", "Excel Files (*.xlsx);;All Files (*)"
        )
        if not file_path:
            return

        try:
            SpreadsheetExportHandler.export_to_excel(self.spreadsheet, file_path)
            QMessageBox.information(
                parent,
                "Success",
                f"Table exported successfully to:\n{file_path}",
            )
        except Exception as e:
            QMessageBox.critical(
                parent, "Export Error", f"Failed to export table:\n{str(e)}"
            )

    def export_to_csv(self, parent=None):
        """
        Export table to CSV with file dialog.

        Args:
            parent: Parent widget for dialogs
        """
        file_path, _ = QFileDialog.getSaveFileName(
            parent, "Export Table as CSV", "", "CSV Files (*.csv);;All Files (*)"
        )
        if not file_path:
            return

        try:
            SpreadsheetExportHandler.export_to_csv(self.spreadsheet, file_path)
            QMessageBox.information(
                parent,
                "Success",
                f"Table exported successfully to:\n{file_path}",
            )
        except Exception as e:
            QMessageBox.critical(
                parent, "Export Error", f"Failed to export table:\n{str(e)}"
            )

    def export_to_json(self, parent=None):
        """
        Export table to JSON with file dialog.

        Args:
            parent: Parent widget for dialogs
        """
        file_path, _ = QFileDialog.getSaveFileName(
            parent, "Export Table as JSON", "", "JSON Files (*.json);;All Files (*)"
        )
        if not file_path:
            return

        try:
            SpreadsheetExportHandler.export_to_json(self.spreadsheet, file_path)
            QMessageBox.information(
                parent,
                "Success",
                f"Table exported successfully to:\n{file_path}",
            )
        except Exception as e:
            QMessageBox.critical(
                parent, "Export Error", f"Failed to export table:\n{str(e)}"
            )


class SpreadsheetExportHandler:
    """
    Low-level export operations - performs actual file I/O and formatting.
    Static methods that work directly with Spreadsheet instances.
    """

    @staticmethod
    def export_to_csv(spreadsheet, file_path):
        """
        Export table data to CSV with formatting columns (no column letters A,B,C).

        Args:
            spreadsheet: Spreadsheet instance to export from
            file_path: Destination file path

        Raises:
            Exception: If export fails
        """
        try:
            with open(file_path, "w", encoding="utf-8", newline="") as f:
                fieldnames = [
                    "Row",
                    "Column",
                    "Value",
                    "FontBold",
                    "FontItalic",
                    "FontUnderline",
                    "TextColor",
                    "BGColor",
                ]
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()

                for r in range(spreadsheet.rowCount()):
                    for c in range(spreadsheet.columnCount()):
                        item = spreadsheet.item(r, c)
                        if not item:
                            continue

                        cell_data = item.get_data()
                        if not cell_data.get("value"):
                            continue

                        font_data = cell_data.get("font", {})
                        row_dict = {
                            "Row": r,
                            "Column": c,
                            "Value": str(cell_data.get("value", "")),
                            "FontBold": font_data.get("bold", False),
                            "FontItalic": font_data.get("italic", False),
                            "FontUnderline": font_data.get("underline", False),
                            "TextColor": cell_data.get("text_color", ""),
                            "BGColor": cell_data.get("bg_color", ""),
                        }
                        writer.writerow(row_dict)
        except Exception as e:
            raise Exception(f"Failed to export to CSV: {str(e)}")

    @staticmethod
    def export_to_json(spreadsheet, file_path):
        """
        Export table data to JSON with full formatting (no column letters A,B,C).

        Args:
            spreadsheet: Spreadsheet instance to export from
            file_path: Destination file path

        Raises:
            Exception: If export fails
        """
        try:
            data = []
            for r in range(spreadsheet.rowCount()):
                row_d = []
                for c in range(spreadsheet.columnCount()):
                    item = spreadsheet.item(r, c)
                    cell_data = item.get_data() if item else {"value": ""}
                    row_d.append(cell_data)
                data.append(row_d)

            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(
                    {"table_data": data}, f, indent=2, ensure_ascii=False
                )
        except Exception as e:
            raise Exception(f"Failed to export to JSON: {str(e)}")

    @staticmethod
    def export_to_excel(spreadsheet, file_path):
        """
        Export table data to Excel with native cell formatting (colors, fonts, styles).

        Args:
            spreadsheet: Spreadsheet instance to export from
            file_path: Destination file path

        Raises:
            Exception: If openpyxl is not available or export fails
        """
        if not OPENPYXL_AVAILABLE:
            raise Exception(
                "openpyxl library is not installed. Install with: pip install openpyxl"
            )

        try:
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = (
                f"Comment{spreadsheet.comment_number}"
                if spreadsheet.comment_number
                else "Data"
            )

            # Write data with formatting
            for r in range(spreadsheet.rowCount()):
                for c in range(spreadsheet.columnCount()):
                    item = spreadsheet.item(r, c)
                    if not item:
                        continue

                    cell_data = item.get_data()
                    if not cell_data.get("value"):
                        continue

                    # Excel cell (1-indexed)
                    excel_cell = ws.cell(row=r + 1, column=c + 1)

                    # Set cell value
                    excel_cell.value = str(cell_data.get("value", ""))

                    # Apply font formatting
                    font_data = cell_data.get("font", {})
                    text_color = cell_data.get("text_color", None)

                    font_kwargs = {
                        "bold": font_data.get("bold", False),
                        "italic": font_data.get("italic", False),
                        "underline": "single"
                        if font_data.get("underline", False)
                        else None,
                    }

                    # Add text color if specified
                    if text_color and text_color.startswith("#"):
                        # Remove # and use as RGB
                        font_kwargs["color"] = text_color[1:].upper()

                    excel_cell.font = Font(
                        **{k: v for k, v in font_kwargs.items() if v is not None}
                    )

                    # Apply background color
                    bg_color = cell_data.get("bg_color", None)
                    if bg_color and bg_color.startswith("#"):
                        # Remove # for Excel color format
                        color_code = bg_color[1:].upper()
                        excel_cell.fill = PatternFill(
                            start_color=color_code,
                            end_color=color_code,
                            fill_type="solid",
                        )

                    # Left align and wrap text
                    excel_cell.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )

            # Auto-adjust column widths
            for c in range(spreadsheet.columnCount()):
                max_length = 10
                column_letter = chr(65 + (c % 26))
                for r in range(spreadsheet.rowCount()):
                    cell = ws.cell(row=r + 1, column=c + 1)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            # Save workbook
            wb.save(file_path)

        except Exception as e:
            raise Exception(f"Failed to export to Excel: {str(e)}")
